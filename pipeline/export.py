"""Backend-only export of the latest successful snapshot per retailer to .xlsx.

No LLM/AI involvement — pure SQL reads against the existing snapshots table
plus openpyxl writes. Runs the same on the first execution as on the
thousandth.
"""

import os
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

import config

_HEADER_FONT = Font(bold=True)


def _latest_successful_run_id(conn: sqlite3.Connection, retailer: str) -> str | None:
    row = conn.execute(
        "SELECT run_id FROM run_log WHERE retailer = ? AND status = 'success' "
        "ORDER BY finished_at DESC LIMIT 1",
        (retailer,),
    ).fetchone()
    return row[0] if row else None


def _snapshot_rows(conn: sqlite3.Connection, retailer: str, run_id: str) -> list[sqlite3.Row]:
    conn.row_factory = sqlite3.Row
    cur = conn.execute(
        "SELECT * FROM snapshots WHERE retailer = ? AND run_id = ?",
        (retailer, run_id),
    )
    return cur.fetchall()


def _write_sheet(ws, headers: list[str], rows: list[tuple]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)


def export_latest_snapshot(conn: sqlite3.Connection, output_path: Path) -> None:
    conn.row_factory = sqlite3.Row
    per_retailer_headers = ["snapshot_datetime"] + config.COLUMNS
    combined_headers = ["retailer"] + per_retailer_headers

    wb = Workbook()
    wb.remove(wb.active)

    combined_rows: list[tuple] = []
    for retailer in config.RETAILERS:
        name = retailer["name"]
        run_id = _latest_successful_run_id(conn, name)
        rows = _snapshot_rows(conn, name, run_id) if run_id else []

        sheet_rows = [tuple(row[col] for col in per_retailer_headers) for row in rows]
        ws = wb.create_sheet(title=name)
        _write_sheet(ws, per_retailer_headers, sheet_rows)

        combined_rows.extend((name,) + row for row in sheet_rows)

    ws_all = wb.create_sheet(title="Todos")
    _write_sheet(ws_all, combined_headers, combined_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)
    try:
        os.replace(tmp_path, output_path)
    except PermissionError:
        # If file is locked (open in Excel), try to save to alternate location
        backup_path = output_path.with_stem(f"{output_path.stem}_latest_backup")
        try:
            os.replace(tmp_path, backup_path)
        except Exception:
            pass
        raise
