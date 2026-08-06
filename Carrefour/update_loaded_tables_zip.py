import csv
import io
import re
import shutil
import zipfile
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


WORKBOOK = Path("Prueba Scrapping Mercadona - vfinal.xlsx")
CSV_PATH = Path("carrefour_modelo_comun.csv")
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)


def col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def cell(col_idx, row_idx, value, style=None):
    c = ET.Element(f"{{{NS_MAIN}}}c", {"r": f"{col_letter(col_idx)}{row_idx}"})
    if style:
        c.set("s", style)
    if value is None or value == "":
        return c
    text = str(value)
    low = text.lower()
    if low in ("true", "false"):
        c.set("t", "b")
        v = ET.SubElement(c, f"{{{NS_MAIN}}}v")
        v.text = "1" if low == "true" else "0"
        return c
    if re.fullmatch(r"-?\d+(\.\d+)?", text):
        c.set("t", "n")
        v = ET.SubElement(c, f"{{{NS_MAIN}}}v")
        v.text = text
        return c
    c.set("t", "inlineStr")
    is_el = ET.SubElement(c, f"{{{NS_MAIN}}}is")
    t_el = ET.SubElement(is_el, f"{{{NS_MAIN}}}t")
    if text.strip() != text:
        t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    t_el.text = text
    return c


def load_csv_rows():
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        return reader.fieldnames, list(reader)


def current_table_data():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    ws = wb["Retailers_Final"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v) if v is not None else "" for v in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1:]]
    non_carrefour = [row for row in data if row.get("retailer") != "Carrefour"]
    return headers, non_carrefour


def workbook_sheet_targets(zip_entries):
    workbook = ET.fromstring(zip_entries["xl/workbook.xml"])
    rels = ET.fromstring(zip_entries["xl/_rels/workbook.xml.rels"])
    rid_to_target = {
        rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
        for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship")
    }
    mapping = {}
    for sheet in workbook.findall(f".//{{{NS_MAIN}}}sheet"):
        rid = sheet.attrib[f"{{{NS_REL}}}id"]
        target = rid_to_target[rid]
        if not target.startswith("xl/"):
            target = "xl/" + target
        mapping[sheet.attrib["name"]] = target
    return mapping


def table_target_for_sheet(zip_entries, sheet_path):
    rel_path = "xl/worksheets/_rels/" + Path(sheet_path).name + ".rels"
    if rel_path not in zip_entries:
        return None
    rels = ET.fromstring(zip_entries[rel_path])
    for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship"):
        if "table" in rel.attrib.get("Type", ""):
            target = rel.attrib["Target"]
            if target.startswith("../"):
                return "xl/" + target[3:]
            return str(Path(sheet_path).parent / target).replace("\\", "/")
    return None


def existing_styles(sheet_xml):
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(f"{{{NS_MAIN}}}sheetData")
    header_styles = {}
    data_styles = {}
    if sheet_data is None:
        return header_styles, data_styles
    for row in sheet_data.findall(f"{{{NS_MAIN}}}row"):
        row_idx = int(row.attrib["r"])
        for c in row.findall(f"{{{NS_MAIN}}}c"):
            ref = c.attrib.get("r", "")
            col = re.sub(r"\d+", "", ref)
            style = c.attrib.get("s")
            idx = sum((ord(ch) - 64) * (26 ** i) for i, ch in enumerate(reversed(col)))
            if row_idx == 1 and style:
                header_styles[idx] = style
            elif row_idx == 2 and style:
                data_styles[idx] = style
    return header_styles, data_styles


def patch_sheet(zip_entries, sheet_path, table_path, headers, rows):
    root = ET.fromstring(zip_entries[sheet_path])
    sheet_data = root.find(f"{{{NS_MAIN}}}sheetData")
    if sheet_data is None:
        sheet_data = ET.SubElement(root, f"{{{NS_MAIN}}}sheetData")
    header_styles, data_styles = existing_styles(zip_entries[sheet_path])
    for child in list(sheet_data):
        sheet_data.remove(child)
    all_rows = [headers] + [[row.get(h, "") for h in headers] for row in rows]
    for r_idx, values in enumerate(all_rows, start=1):
        row_el = ET.Element(f"{{{NS_MAIN}}}row", {"r": str(r_idx), "spans": f"1:{len(headers)}"})
        for c_idx, value in enumerate(values, start=1):
            style = header_styles.get(c_idx) if r_idx == 1 else data_styles.get(c_idx)
            row_el.append(cell(c_idx, r_idx, value, style))
        sheet_data.append(row_el)
    ref = f"A1:{col_letter(len(headers))}{len(all_rows)}"
    dimension = root.find(f"{{{NS_MAIN}}}dimension")
    if dimension is None:
        dimension = ET.Element(f"{{{NS_MAIN}}}dimension")
        root.insert(0, dimension)
    dimension.set("ref", ref)
    zip_entries[sheet_path] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    if table_path and table_path in zip_entries:
        table = ET.fromstring(zip_entries[table_path])
        table.set("ref", ref)
        table.set("totalsRowCount", "0")
        auto_filter = table.find(f"{{{NS_MAIN}}}autoFilter")
        if auto_filter is not None:
            auto_filter.set("ref", ref)
        zip_entries[table_path] = ET.tostring(table, encoding="utf-8", xml_declaration=True)


def main():
    csv_headers, carrefour_rows = load_csv_rows()
    retail_headers, non_carrefour = current_table_data()
    retail_rows = []
    for row in non_carrefour:
        retail_rows.append({h: row.get(h, "") for h in retail_headers})
    for row in carrefour_rows:
        retail_rows.append({h: row.get(h, "") for h in retail_headers})

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = WORKBOOK.with_name(f"{WORKBOOK.stem}.backup_loaded_tables_{timestamp}{WORKBOOK.suffix}")
    shutil.copy2(WORKBOOK, backup)

    with zipfile.ZipFile(WORKBOOK, "r") as zin:
        infos = zin.infolist()
        zip_entries = {info.filename: zin.read(info.filename) for info in infos}

    targets = workbook_sheet_targets(zip_entries)
    for sheet_name, headers, rows in [
        ("Carrefour_ModeloComun", csv_headers, carrefour_rows),
        ("Retailers_Final", retail_headers, retail_rows),
    ]:
        sheet_path = targets[sheet_name]
        table_path = table_target_for_sheet(zip_entries, sheet_path)
        patch_sheet(zip_entries, sheet_path, table_path, headers, rows)

    with zipfile.ZipFile(WORKBOOK, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for info in infos:
            data = zip_entries[info.filename]
            new_info = zipfile.ZipInfo(info.filename, date_time=info.date_time)
            new_info.comment = info.comment
            new_info.extra = info.extra
            new_info.internal_attr = info.internal_attr
            new_info.external_attr = info.external_attr
            new_info.create_system = info.create_system
            new_info.compress_type = zipfile.ZIP_DEFLATED
            zout.writestr(new_info, data)

    print(f"backup={backup}")
    print(f"carrefour_rows={len(carrefour_rows)}")
    print(f"retailers_final_rows={len(retail_rows)}")
    print(f"retailers_final_carrefour_rows={sum(1 for row in retail_rows if row.get('retailer') == 'Carrefour')}")


if __name__ == "__main__":
    main()
