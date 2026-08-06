from collections import Counter

from openpyxl import load_workbook


WORKBOOK = "Prueba Scrapping Mercadona - vfinal.xlsx"


def count_retailers(sheet_name: str) -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    retailer_col = headers.index("retailer") + 1
    values = [
        row[0]
        for row in ws.iter_rows(
            min_row=2,
            min_col=retailer_col,
            max_col=retailer_col,
            values_only=True,
        )
    ]
    counts = Counter(values)
    print(f"{sheet_name}: data_rows={len(values)}")
    print(f"{sheet_name}: carrefour={counts.get('Carrefour', 0)}")


if __name__ == "__main__":
    count_retailers("Carrefour_ModeloComun")
    count_retailers("Retailers_Final")
